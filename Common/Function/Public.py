# coding=utf-8

import openpyxl
import csv
import pandas
import chardet
import random
from datetime import timedelta, date
import shutil
import os
import yaml

class Public(object):
    """
    说明：
        提供一些常用功能
        1. readExcel():
            读取Excel单元格数据
        2. writeExcel
            写入Excel单元格数据
        3. readCsv():
            读取Csv单元格数据
        4. writeCsv()
            写入Csv单元格数据
        5. getFileCoding()
            获取文件的编码格式
    """

    def __returnCloumn(self, ws, p_column):
        """
        :param ws: sheet页 type=openpyxl.worksheet.worksheet.Worksheet
        :param p_column: 列名或列数 type=int or str
        :return: 列数转换为字母形式
        """
        if isinstance(p_column, (int)):
            return openpyxl.utils.get_column_letter(p_column)  # 数字转换为字母
        elif isinstance(p_column, (str)):
            row_list = list(map(lambda x: x.value, list(ws[1])))  # 解析第一行
            if row_list.count(p_column) == 1:  # 相同字段名存在的个数
                for index, cell in enumerate(row_list):
                    if cell == p_column:
                        return openpyxl.utils.get_column_letter(index + 1)
            elif row_list.count(p_column) == 0:
                raise ValueError("当前sheet页“%s”不存在字段名为“%s”" % (ws.title, p_column))
            else:
                raise ValueError("当前sheet页“%s”存在多个字段名为“%s”" % (ws.title, p_column))
        else:
            raise TypeError(
                "column参数类型应为int或str类型而不是“%s”，int表示第几列，str表示第一行列字段名"
                % p_column.__class__
            )

    def readExcel(self, path=str, column=1, row=1, sheet=1):
        """
        读取Excel某一单元格数据
        :param path: Excel路径 type=str
        :param sheet: 第几个sheet页或sheet页名称 type=int or str
        :param row: 行数 type=int
        :param column: 列数或列字段名 type=int or str
        :return: 单元格的值
        """
        if not isinstance(row, (int)):
            raise TypeError("row参数类型应为int类型而不是“%s”" % row.__class__)
        wb = openpyxl.load_workbook(path)
        if isinstance(sheet, (str)):
            ws = wb.get_sheet_by_name(sheet)
            column = self.__returnCloumn(ws=ws, p_column=column)
            return ws[column + str(row)].value
        elif isinstance(sheet, (int)):
            ws = wb.get_sheet_by_name(wb.sheetnames[sheet - 1])
            column = self.__returnCloumn(ws=ws, p_column=column)
            return ws[column + str(row)].value
        else:
            raise TypeError(
                "sheet参数类型应为int或str类型而不是“%s”，数字表示为第几个sheet页，字符表示sheet页名称"
                % sheet.__class__
            )

    def writeExcel(self, path, value="", column=1, row=1, sheet=1):
        """
        写入Excel指定单元格数据
        :param path: Excel路径 type=str
        :param sheet: 第几个sheet页或sheet页名称 type=int or str
        :param row: 行数 type=int
        :param column: 列数或列字段名 type=int or str
        :param value: 需要写入的数据 type=str
        """
        if not isinstance(row, (int)):
            raise TypeError("row参数类型应为int类型而不是“%s”" % row.__class__)
        if not isinstance(value, (str)):
            raise TypeError("value参数类型应为str类型而不是“%s”" % value.__class__)
        wb = openpyxl.load_workbook(path)
        if isinstance(sheet, (str)):
            ws = wb.get_sheet_by_name(sheet)
            column = self.__returnCloumn(ws=ws, p_column=column)
            ws[column + str(row)].value = value
            wb.save(path)
        elif isinstance(sheet, (int)):
            ws = wb.get_sheet_by_name(wb.sheetnames[sheet - 1])
            column = self.__returnCloumn(ws=ws, p_column=column)
            ws[column + str(row)].value = value
            wb.save(path)
        else:
            raise TypeError(
                "sheet参数类型应为int或str类型而不是“%s”，数字表示为第几个sheet页，字符表示sheet页名称"
                % sheet.__class__
            )

    def readCsv(self, path=str, column=1, row=1, fields=True):
        """
        说明：读取csv某一单元格数据
        注意：文件编码建议使用utf-8
        :param path: csv文件路径 type=str
        :param row: 行数 type=int
        :param column: 字段名或列数 type=str or int
        :param fields: csv文件第一行是否是字段
        :return: 单元格值
        """
        if not isinstance(row, (int)):
            raise TypeError("row参数类型应为int类型而不是“%s”" % row.__class__)
        with open(path, "r", encoding="utf-8") as csv_file:  # 转为list
            csv_reader = csv.reader(csv_file)
            csv_list = list(csv_reader)
        with open(path, "r", encoding="utf-8") as csv_file:  # 转为dict
            csv_dict = csv.DictReader(csv_file)
            csv_dict_list = list(csv_dict)
        if fields:
            if isinstance(column, (str)):
                return csv_dict_list[row - 2][column]
            elif isinstance(column, (int)):
                return csv_list[row - 1][column - 1]
            else:
                raise TypeError(
                    "csv中有字段时，可以使用字段名表示列名类型为str，当然也可以使用列数表示类型为int，但当前类型为“%s”"
                    % column.__class__
                )
        else:
            if not isinstance(column, (int)):
                raise TypeError("csv中没有字段，请使用int类型表示列数，当前类型为“%s”" % column.__class__)
            else:
                return csv_list[row - 1][column - 1]

    def writeCsv(self, path=str, value="", column=1, row=1, fields=True):
        """
        说明：写入csv指定单元格数据
        注意：文件编码建议使用utf-8
        :param path: csv文件路径 type=str
        :param value: 需要写入的数据 type=str
        :param row: 行数 type=int
        :param column: 字段名或列数 type=str or int
        :param fields: csv文件第一行是否是字段
        :return: None
        """
        if not isinstance(row, (int)):
            raise TypeError("row参数类型应为int类型而不是“%s”" % row.__class__)
        if row == 1:
            raise ValueError("第一行为字段行，不允许修改...")
        df = pandas.read_csv(path)
        if fields:
            if isinstance(column, (str)):
                df.at[row - 2, column] = value  # [行数，列名]
            elif isinstance(column, (int)):
                df.iat[row - 2, column - 1] = value  # [行数，列数]
            else:
                raise TypeError(
                    "csv中有字段时，可以使用字段名表示列名类型为str，当然也可以使用列数表示类型为int，但当前类型为“%s”"
                    % column.__class__
                )
        else:
            if not isinstance(column, (int)):
                raise TypeError("csv中没有字段，请使用int类型表示列数，当前类型为“%s”" % column.__class__)
            else:
                df.iat[row - 2, column - 1] = value
        df.to_csv(path_or_buf=path, index=False)

    def getFileCoding(self, filepath=str):
        """
        说明：
            获取文件的编码格式
        :param filepath: 文件路径 type=str
        :return: {'confidence': 1.0, 'encoding': 'UTF-8'} type=dict
        """
        with open(file=filepath, mode="rb") as f:
            s = f.read()
        return chardet.detect(s)

    def randomPhone(self):
        """
        说明:
            生成中国大陆手机号
        :return: str
        """
        prelist = ["130","131","132", "133","134","135","136","137","138","139","147","150","151","152","153","155","156","157","158","159","186","187","188"]
        return random.choice(prelist) + "".join(
            random.choice("0123456789") for i in range(8)
        )

    def randomIdCard(self):
        """
        说明:
            生成中国大陆身份证号
        :return: str
        """
        curpath = os.path.dirname(os.path.realpath(__file__))
        yamlpath = curpath + "\Date\Address.yaml"
        with open(yamlpath, "r", encoding='utf-8') as f:
            cfg = f.read()
        area_dict= yaml.load(cfg)
        id_code_list = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
        check_code_list = [1, 0, "X", 9, 8, 7, 6, 5, 4, 3, 2]
        area_code = random.choice(list(area_dict.keys()))
        age = random.randint(1, 99)  # 年龄
        gender = random.randint(0, 1)  # 0-女性 1-男性
        datestring = str(
            date(date.today().year - age, 1, 1) + timedelta(days=random.randint(0, 364))
        ).replace("-", "")
        rd = random.randint(0, 999)
        if gender == 0:
            gender_num = rd if rd % 2 == 0 else rd + 1
        else:
            gender_num = rd if rd % 2 == 1 else rd - 1
        result = str(area_code) + datestring + str(gender_num).zfill(3)
        return result + str(
            check_code_list[
                sum([a * b for a, b in zip(id_code_list, [int(a) for a in result])])
                % 11
            ]
        )

    def RandomEmail(self, emailType=None, rang=None):
        """
        说明:
            生成邮箱
        :return: str
        """
        __emailtype = [
            "@qq.com",
            "@163.com",
            "@126.com",
            "@outlook.com",
            "@gmail.com",
            "@yahoo.com",
            "@sina.com",
            "@sohu.com",
        ]
        # 如果没有指定邮箱类型，默认在 __emailtype中随机一个
        if emailType == None:
            __randomEmail = random.choice(__emailtype)
        else:
            __randomEmail = emailType
        # 如果没有指定邮箱长度，默认在4-10之间随机
        if rang == None:
            __rang = random.randint(4, 10)
        else:
            __rang = int(rang)
        __Number = "0123456789qbcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPWRSTUVWXYZ"
        __randomNumber = "".join(random.choice(__Number) for i in range(__rang))
        _email = __randomNumber + __randomEmail
        return _email

    def randomIPAdress(self):
        """
        说明:
            生成IP地址
        :return: str
        """
        return ".".join(str(random.randint(1, 254)) for i in range(4))

    def make_archive(self, base_name=str, format="zip", root_dir=None):
        """
        说明：
            压缩指定文件夹\目录
        :param base_name: 压缩后文件名\路径
        :param format: 压缩格式 "zip", "tar", "bztar", "gztar"
        :param root_dir: 被压缩目录
        :return: 压缩后文件路径
        """
        return shutil.make_archive(
                    base_name=base_name,
                    format=format,
                    root_dir=root_dir
               )

    def unpack_archive(self, filename=str, extract_dir=None, format=None):
        """
        说明：
            解压指定文件
        :param filename: 待解压文件名\路径
        :param extract_dir: 解压到指定目录
        :param format: 解压格式，如果未指定则取压缩文件的后缀名格式
        """
        shutil.unpack_archive(
            filename=filename,
            extract_dir=extract_dir,
            format=format
        )
if __name__ == '__main__':
    print(Public().randomIdCard())
    print(Public().randomPhone())